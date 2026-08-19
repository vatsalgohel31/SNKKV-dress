import io
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import StandardPrice, StudentDressEntry, STANDARD_CHOICES, MEDIUM_CHOICES, PACKAGE_CHOICES, SECTION_CHOICES
from .forms import StudentDressEntryForm


def is_admin_user(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def login_view(request):
    if request.user.is_authenticated:
        if is_admin_user(request.user):
            return redirect('entry')
        return redirect('viewer')
        
    if request.method == 'POST':
        u = request.POST.get('username', '').strip()
        p = request.POST.get('password', '').strip()
        user = authenticate(request, username=u, password=p)
        if user is not None:
            login(request, user)
            messages.success(request, f"Welcome, {user.username}!")
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            if is_admin_user(user):
                return redirect('entry')
            return redirect('viewer')
        else:
            messages.error(request, "Invalid username or password. Please try again.")
            
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out successfully.")
    return redirect('login')


@login_required
def entry_view(request):
    # View-only protection
    if not is_admin_user(request.user):
        messages.error(request, "Access restricted: View-only users cannot add student records.")
        return redirect('viewer')

    # Ensure default standard prices exist (1..12)
    for std in range(1, 13):
        StandardPrice.objects.get_or_create(
            standard=std,
            defaults={
                'dress_price': Decimal('350.00'),
                'extra_dress_price': Decimal('300.00'),
                'dupatta_price': Decimal('120.00'),
                'extra_dupatta_price': Decimal('100.00')
            }
        )

    if request.method == 'POST':
        form = StudentDressEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            
            # Sync booleans and recalculate if needed
            entry.has_dress = 'has_dress' in request.POST
            entry.has_extra_dress = 'has_extra_dress' in request.POST
            entry.has_dupatta = 'has_dupatta' in request.POST
            entry.has_extra_dupatta = 'has_extra_dupatta' in request.POST
            
            # If price was provided from UI or compute from DB
            sub_total = request.POST.get('total_price')
            if sub_total:
                try:
                    entry.total_price = Decimal(str(sub_total))
                except Exception:
                    entry.total_price = entry.compute_price_from_standard()
            else:
                entry.total_price = entry.compute_price_from_standard()
                
            entry.save()
            messages.success(request, f"Entry for '{entry.name}' (Std {entry.standard}) saved successfully! Total: ₹{entry.total_price}")
            
            action = request.POST.get('submit_action', 'save_new')
            if action == 'save_view':
                return redirect('viewer')
            return redirect('entry')
        else:
            messages.error(request, "Please check the form for errors.")
    else:
        form = StudentDressEntryForm()

    recent_entries = StudentDressEntry.objects.all()[:5]
    standards = STANDARD_CHOICES
    mediums = MEDIUM_CHOICES
    packages = PACKAGE_CHOICES

    # Get standard prices as json dictionary for fast frontend calculation
    prices_map = {
        sp.standard: {
            'dress': float(sp.dress_price),
            'extra_dress': float(sp.extra_dress_price),
            'dupatta': float(sp.dupatta_price),
            'extra_dupatta': float(sp.extra_dupatta_price)
        } for sp in StandardPrice.objects.all()
    }

    return render(request, 'entry.html', {
        'form': form,
        'recent_entries': recent_entries,
        'standards': standards,
        'mediums': mediums,
        'packages': packages,
        'prices_map': prices_map,
    })


@login_required
def pricing_view(request):
    """Dedicated Standard-wise price configuration matrix for Std 1 to 12."""
    if not is_admin_user(request.user):
        messages.error(request, "Access restricted: View-only users cannot modify pricing.")
        return redirect('viewer')

    for std in range(1, 13):
        StandardPrice.objects.get_or_create(
            standard=std,
            defaults={
                'dress_price': Decimal('350.00'),
                'extra_dress_price': Decimal('300.00'),
                'dupatta_price': Decimal('120.00'),
                'extra_dupatta_price': Decimal('100.00')
            }
        )

    if request.method == 'POST':
        # Batch update all 12 standards
        updated_count = 0
        for std in range(1, 13):
            d_p = request.POST.get(f'dress_price_{std}')
            ed_p = request.POST.get(f'extra_dress_price_{std}')
            dup_p = request.POST.get(f'dupatta_price_{std}')
            edup_p = request.POST.get(f'extra_dupatta_price_{std}')

            if d_p is not None:
                sp, _ = StandardPrice.objects.get_or_create(standard=std)
                sp.dress_price = Decimal(d_p or '0.00')
                sp.extra_dress_price = Decimal(ed_p or '0.00')
                sp.dupatta_price = Decimal(dup_p or '0.00')
                sp.extra_dupatta_price = Decimal(edup_p or '0.00')
                sp.save()
                updated_count += 1
                
        messages.success(request, f"Standard-wise prices updated successfully for {updated_count} standards!")
        return redirect('pricing')

    standard_prices = StandardPrice.objects.all().order_by('standard')
    return render(request, 'pricing.html', {
        'standard_prices': standard_prices,
    })


@login_required
def viewer_view(request):
    """Excel-like data viewer page with multi-column filtering, search, metrics, and right-side PDF button."""
    entries = StudentDressEntry.objects.all()

    # Query params for filtering
    search_q = request.GET.get('search', '').strip()
    medium_filter = request.GET.get('medium', '').strip()
    section_filter = request.GET.get('section', '').strip()
    std_filter = request.GET.get('standard', '').strip()
    package_filter = request.GET.get('package', '').strip()
    dress_filter = request.GET.get('has_dress', '').strip()
    dupatta_filter = request.GET.get('has_dupatta', '').strip()
    extra_dress_filter = request.GET.get('has_extra_dress', '').strip()
    extra_dupatta_filter = request.GET.get('has_extra_dupatta', '').strip()

    if search_q:
        entries = entries.filter(name__icontains=search_q)
    if medium_filter:
        entries = entries.filter(medium=medium_filter)
    if section_filter == 'PRIMARY':
        entries = entries.filter(standard__gte=1, standard__lte=8)
    elif section_filter == 'SECONDARY':
        entries = entries.filter(standard__in=[9, 10])
    elif section_filter == 'HIGHER_SECONDARY':
        entries = entries.filter(standard__in=[11, 12])
    if std_filter:
        try:
            entries = entries.filter(standard=int(std_filter))
        except ValueError:
            pass
    if package_filter:
        entries = entries.filter(package_model=package_filter)
    if dress_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_dress=True)
    if dupatta_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_dupatta=True)
    if extra_dress_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_extra_dress=True)
    if extra_dupatta_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_extra_dupatta=True)

    # Aggregates for top dashboard cards
    total_records = entries.count()
    total_amount = entries.aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')
    total_dresses = entries.aggregate(Sum('dress_qty'))['dress_qty__sum'] or 0
    total_dupattas = entries.aggregate(Sum('dupatta_qty'))['dupatta_qty__sum'] or 0
    gujarati_count = entries.filter(medium='GUJARATI').count()
    english_count = entries.filter(medium='ENGLISH').count()

    primary_count = entries.filter(standard__gte=1, standard__lte=8).count()
    secondary_count = entries.filter(standard__in=[9, 10]).count()
    higher_sec_count = entries.filter(standard__in=[11, 12]).count()

    standards = STANDARD_CHOICES
    mediums = MEDIUM_CHOICES
    packages = PACKAGE_CHOICES
    sections = SECTION_CHOICES

    return render(request, 'viewer.html', {
        'entries': entries,
        'total_records': total_records,
        'total_amount': total_amount,
        'total_dresses': total_dresses,
        'total_dupattas': total_dupattas,
        'gujarati_count': gujarati_count,
        'english_count': english_count,
        'primary_count': primary_count,
        'secondary_count': secondary_count,
        'higher_sec_count': higher_sec_count,
        'standards': standards,
        'mediums': mediums,
        'packages': packages,
        'sections': sections,
        'current_search': search_q,
        'current_medium': medium_filter,
        'current_section': section_filter,
        'current_std': std_filter,
        'current_package': package_filter,
        'current_has_dress': dress_filter,
        'current_has_dupatta': dupatta_filter,
        'current_has_extra_dress': extra_dress_filter,
        'current_has_extra_dupatta': extra_dupatta_filter,
        'is_admin': is_admin_user(request.user),
    })


@login_required
def edit_entry_view(request, pk):
    if not is_admin_user(request.user):
        messages.error(request, "Access restricted: View-only users cannot edit student records.")
        return redirect('viewer')

    entry = get_object_or_404(StudentDressEntry, pk=pk)
    if request.method == 'POST':
        form = StudentDressEntryForm(request.POST, instance=entry)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.has_dress = 'has_dress' in request.POST
            updated.has_extra_dress = 'has_extra_dress' in request.POST
            updated.has_dupatta = 'has_dupatta' in request.POST
            updated.has_extra_dupatta = 'has_extra_dupatta' in request.POST
            
            sub_total = request.POST.get('total_price')
            if sub_total:
                try:
                    updated.total_price = Decimal(str(sub_total))
                except Exception:
                    updated.total_price = updated.compute_price_from_standard()
            else:
                updated.total_price = updated.compute_price_from_standard()
                
            updated.save()
            messages.success(request, f"Entry for {updated.name} updated successfully!")
            return redirect('viewer')
    else:
        form = StudentDressEntryForm(instance=entry)

    prices_map = {
        sp.standard: {
            'dress': float(sp.dress_price),
            'extra_dress': float(sp.extra_dress_price),
            'dupatta': float(sp.dupatta_price),
            'extra_dupatta': float(sp.extra_dupatta_price)
        } for sp in StandardPrice.objects.all()
    }

    return render(request, 'entry.html', {
        'form': form,
        'is_edit': True,
        'entry': entry,
        'standards': STANDARD_CHOICES,
        'mediums': MEDIUM_CHOICES,
        'packages': PACKAGE_CHOICES,
        'prices_map': prices_map,
    })


@login_required
def delete_entry_view(request, pk):
    if not is_admin_user(request.user):
        messages.error(request, "Access restricted: View-only users cannot delete records.")
        return redirect('viewer')

    entry = get_object_or_404(StudentDressEntry, pk=pk)
    if request.method == 'POST':
        name = entry.name
        entry.delete()
        messages.success(request, f"Record for '{name}' was deleted.")
        return redirect('viewer')
    return redirect('viewer')


@login_required
def get_standard_price_api(request, std):
    """Returns JSON price info for given standard."""
    try:
        sp = StandardPrice.objects.get(standard=int(std))
        return JsonResponse({
            'success': True,
            'standard': sp.standard,
            'dress_price': float(sp.dress_price),
            'extra_dress_price': float(sp.extra_dress_price),
            'dupatta_price': float(sp.dupatta_price),
            'extra_dupatta_price': float(sp.extra_dupatta_price),
        })
    except StandardPrice.DoesNotExist:
        return JsonResponse({
            'success': False,
            'dress_price': 350.0,
            'extra_dress_price': 300.0,
            'dupatta_price': 120.0,
            'extra_dupatta_price': 100.0,
        })


@login_required
def export_pdf_view(request):
    """Generates PDF of filtered or all student dress entries."""
    entries = StudentDressEntry.objects.all()

    search_q = request.GET.get('search', '').strip()
    medium_filter = request.GET.get('medium', '').strip()
    section_filter = request.GET.get('section', '').strip()
    std_filter = request.GET.get('standard', '').strip()
    package_filter = request.GET.get('package', '').strip()
    dress_filter = request.GET.get('has_dress', '').strip()
    dupatta_filter = request.GET.get('has_dupatta', '').strip()
    extra_dress_filter = request.GET.get('has_extra_dress', '').strip()
    extra_dupatta_filter = request.GET.get('has_extra_dupatta', '').strip()

    filter_desc = []
    if search_q:
        entries = entries.filter(name__icontains=search_q)
        filter_desc.append(f"Search: '{search_q}'")
    if medium_filter:
        entries = entries.filter(medium=medium_filter)
        filter_desc.append(f"Medium: {medium_filter.capitalize()}")
    if section_filter == 'PRIMARY':
        entries = entries.filter(standard__gte=1, standard__lte=8)
        filter_desc.append("Section: Primary (Std 1-8)")
    elif section_filter == 'SECONDARY':
        entries = entries.filter(standard__in=[9, 10])
        filter_desc.append("Section: Secondary (Std 9, 10)")
    elif section_filter == 'HIGHER_SECONDARY':
        entries = entries.filter(standard__in=[11, 12])
        filter_desc.append("Section: Higher Secondary (Std 11, 12)")
    if std_filter:
        try:
            entries = entries.filter(standard=int(std_filter))
            filter_desc.append(f"Standard: Std {std_filter}")
        except ValueError:
            pass
    if package_filter:
        entries = entries.filter(package_model=package_filter)
        filter_desc.append(f"Package: {package_filter}")
    if dress_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_dress=True)
        filter_desc.append("1 Dress: Yes")
    if dupatta_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_dupatta=True)
        filter_desc.append("1 Dupatta: Yes")
    if extra_dress_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_extra_dress=True)
        filter_desc.append("Extra Dress: Yes")
    if extra_dupatta_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_extra_dupatta=True)
        filter_desc.append("Extra Dupatta: Yes")

    filter_text = ", ".join(filter_desc) if filter_desc else "All Records (No Filters Applied)"

    # Create PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'MainTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1e293b'),
        alignment=1
    )
    subtitle_style = ParagraphStyle(
        'SubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#64748b'),
        alignment=1
    )
    filter_label_style = ParagraphStyle(
        'FilterLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#475569'),
        alignment=0
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1e293b')
    )
    cell_bold = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#0f172a')
    )
    badge_yes = ParagraphStyle(
        'BadgeYes',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#166534'),
        alignment=1
    )
    badge_no = ParagraphStyle(
        'BadgeNo',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#94a3b8'),
        alignment=1
    )

    # Document Header
    elements.append(Paragraph("SCHOOL DRESS ORDER &amp; SUMMARY SHEET", title_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%d-%b-%Y %I:%M %p')} | Total Records: {entries.count()}", subtitle_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>Active Filters:</b> {filter_text}", filter_label_style))
    elements.append(Spacer(1, 10))

    # Table Header & Rows (Measurement removed)
    table_data = [
        [
            Paragraph("<b>#</b>", cell_bold),
            Paragraph("<b>Student Name</b>", cell_bold),
            Paragraph("<b>Section</b>", cell_bold),
            Paragraph("<b>Std</b>", cell_bold),
            Paragraph("<b>Medium</b>", cell_bold),
            Paragraph("<b>1 Dress</b>", cell_bold),
            Paragraph("<b>1 Dupatta</b>", cell_bold),
            Paragraph("<b>1 Extra Dress</b>", cell_bold),
            Paragraph("<b>1 Extra Dupatta</b>", cell_bold),
            Paragraph("<b>Dress Qty</b>", cell_bold),
            Paragraph("<b>Dupatta Qty</b>", cell_bold),
            Paragraph("<b>Total Amount</b>", cell_bold),
        ]
    ]

    tot_amt = Decimal('0.00')
    tot_dresses = 0
    tot_dupattas = 0

    for idx, e in enumerate(entries, 1):
        tot_amt += e.total_price
        tot_dresses += e.dress_qty
        tot_dupattas += e.dupatta_qty

        d_1 = Paragraph("YES", badge_yes) if e.has_dress else Paragraph("-", badge_no)
        dup_1 = Paragraph("YES", badge_yes) if e.has_dupatta else Paragraph("-", badge_no)
        ex_d = Paragraph("YES", badge_yes) if e.has_extra_dress else Paragraph("-", badge_no)
        ex_dup = Paragraph("YES", badge_yes) if e.has_extra_dupatta else Paragraph("-", badge_no)

        table_data.append([
            Paragraph(str(idx), cell_style),
            Paragraph(f"<b>{e.name}</b>", cell_style),
            Paragraph(e.section_name, cell_style),
            Paragraph(f"Std {e.standard}", cell_style),
            Paragraph(e.get_medium_display(), cell_style),
            d_1,
            dup_1,
            ex_d,
            ex_dup,
            Paragraph(str(e.dress_qty), cell_style),
            Paragraph(str(e.dupatta_qty), cell_style),
            Paragraph(f"₹{e.total_price:.2f}", cell_bold),
        ])

    # Summary row at bottom
    table_data.append([
        Paragraph("<b>TOTAL</b>", cell_bold),
        Paragraph(f"<b>{entries.count()} Students</b>", cell_bold),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph("", cell_style),
        Paragraph(f"<b>{tot_dresses}</b>", cell_bold),
        Paragraph(f"<b>{tot_dupattas}</b>", cell_bold),
        Paragraph(f"<b>₹{tot_amt:.2f}</b>", cell_bold),
    ])

    col_widths = [26, 170, 85, 50, 75, 55, 60, 70, 75, 55, 65, 95]
    pdf_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.HexColor('#ffffff'), colors.HexColor('#f8fafc')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e2e8f0')),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#0f172a')),
    ])
    pdf_table.setStyle(table_style)
    elements.append(pdf_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"school_dress_report_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def export_excel_view(request):
    """Exports active filtered entries directly into an Excel (.xlsx) spreadsheet."""
    entries = StudentDressEntry.objects.all()

    search_q = request.GET.get('search', '').strip()
    medium_filter = request.GET.get('medium', '').strip()
    section_filter = request.GET.get('section', '').strip()
    std_filter = request.GET.get('standard', '').strip()
    package_filter = request.GET.get('package', '').strip()
    dress_filter = request.GET.get('has_dress', '').strip()
    dupatta_filter = request.GET.get('has_dupatta', '').strip()
    extra_dress_filter = request.GET.get('has_extra_dress', '').strip()
    extra_dupatta_filter = request.GET.get('has_extra_dupatta', '').strip()

    if search_q:
        entries = entries.filter(name__icontains=search_q)
    if medium_filter:
        entries = entries.filter(medium=medium_filter)
    if section_filter == 'PRIMARY':
        entries = entries.filter(standard__gte=1, standard__lte=8)
    elif section_filter == 'SECONDARY':
        entries = entries.filter(standard__in=[9, 10])
    elif section_filter == 'HIGHER_SECONDARY':
        entries = entries.filter(standard__in=[11, 12])
    if std_filter:
        try:
            entries = entries.filter(standard=int(std_filter))
        except ValueError:
            pass
    if package_filter:
        entries = entries.filter(package_model=package_filter)
    if dress_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_dress=True)
    if dupatta_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_dupatta=True)
    if extra_dress_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_extra_dress=True)
    if extra_dupatta_filter in ['1', 'true', 'True']:
        entries = entries.filter(has_extra_dupatta=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "School Dress Orders"

    # Header styling
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True)
    regular_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "Sr No", "Student Name", "Section", "Standard", "Medium",
        "1 Dress", "1 Dupatta", "1 Extra Dress", "1 Extra Dupatta",
        "Dress Qty", "Dupatta Qty", "Total Amount (₹)", "Date Added"
    ]

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tot_amt = Decimal('0.00')
    tot_dresses = 0
    tot_dupattas = 0

    for idx, e in enumerate(entries, 1):
        tot_amt += e.total_price
        tot_dresses += e.dress_qty
        tot_dupattas += e.dupatta_qty

        row = [
            idx,
            e.name,
            e.section_name,
            f"Std {e.standard}",
            e.get_medium_display(),
            "Yes" if e.has_dress else "No",
            "Yes" if e.has_dupatta else "No",
            "Yes" if e.has_extra_dress else "No",
            "Yes" if e.has_extra_dupatta else "No",
            e.dress_qty,
            e.dupatta_qty,
            float(e.total_price),
            e.created_at.strftime("%d-%b-%Y")
        ]
        ws.append(row)
        curr_row = ws.max_row
        for col_num in range(1, len(row) + 1):
            c = ws.cell(row=curr_row, column=col_num)
            c.font = regular_font
            c.border = thin_border
            if col_num in [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13]:
                c.alignment = Alignment(horizontal="center")
            elif col_num == 12:
                c.alignment = Alignment(horizontal="right")

    # Summary Row
    summary_row = [
        "Total", f"{entries.count()} Students", "", "", "",
        "", "", "", "",
        tot_dresses, tot_dupattas, float(tot_amt), ""
    ]
    ws.append(summary_row)
    sum_r = ws.max_row
    for col_num in range(1, len(summary_row) + 1):
        c = ws.cell(row=sum_r, column=col_num)
        c.font = bold_font
        c.border = thin_border
        c.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"school_dress_sheet_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


